"""
台幣 (TWD) 兌主要貨幣匯率 Excel 產生程式
日期：2026/06/05
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


def create_rates_excel():
    """產生台幣匯率 Excel 檔案"""
    
    # 建立工作簿
    wb = openpyxl.Workbook()
    
    # 定義樣式
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Microsoft JhengHei', bold=True, color='FFFFFF', size=11)
    title_font = Font(name='Microsoft JhengHei', bold=True, size=14, color='1F4E79')
    data_font = Font(name='Microsoft JhengHei', size=10)
    data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 輔助函數：寫入標題列
    def write_header(ws, row, headers):
        for col, val in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
    
    # 輔助函數：寫入資料列
    def write_data(ws, start_row, rows, widths=None):
        for ri, rd in enumerate(rows):
            for ci, val in enumerate(rd):
                cell = ws.cell(row=start_row + ri, column=ci + 1, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = border_style
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
    
    # ===== Sheet 1: 匯率總覽 =====
    ws1 = wb.active
    ws1.title = "匯率總覽"
    ws1.merge_cells('A1:H1')
    ws1['A1'] = '台幣 (TWD) 兌主要貨幣匯率總覽 (2026/06/05)'
    ws1['A1'].font = title_font
    ws1.row_dimensions[1].height = 30
    
    write_header(ws1, 3, [
        "幣別", "現匯價格 (1 TWD=)", "7 日最高", "7 日最低",
        "7 日平均", "24h 變化", "7 天變化", "12 個月趨勢"
    ])
    write_data(ws1, 4, [
        ["🇺🇸 美元 (USD)", "$0.0317", "$0.0319", "$0.0317", "$0.0318", "-0.09%", "-0.30%", "相對貶值 5.31%"],
        ["🇯🇵 日圓 (JPY)", "¥5.08", "¥5.095", "¥5.066", "¥5.075", "-0.01%", "+0.18%", "相對貶值 6.12%"],
        ["🇪🇺 歐元 (EUR)", "€0.026", "-", "-", "-", "-", "-", "-"],
        ["🇬🇧 英鎊 (GBP)", "£0.024", "-", "-", "-", "-", "-", "-"],
        ["🇭🇰 港幣 (HKD)", "HK$0.247", "-", "-", "-", "-", "-", "-"],
        ["🇨🇳 人民幣 (CNY)", "¥0.215", "-", "-", "-", "-", "-", "-"],
    ], [14, 20, 14, 14, 14, 14, 14, 18])
    ws1.auto_filter.ref = "A3:H9"
    
    # ===== Sheet 2: 台幣 <-> 美元 =====
    ws2 = wb.create_sheet("台幣 <-> 美元")
    ws2.merge_cells('A1:E1')
    ws2['A1'] = '台幣 (TWD) 兌美元 (USD) 換算對照表'
    ws2['A1'].font = title_font
    
    write_header(ws2, 3, ["台幣金額", "美元金額", "", "美元金額", "台幣金額"])
    write_data(ws2, 4, [
        ["NT$100", "$3.17", "←", "USD $1", "NT$31.52"],
        ["NT$500", "$15.85", "←", "USD $10", "NT$315.20"],
        ["NT$1,000", "$31.70", "←", "USD $100", "NT$3,152"],
        ["NT$5,000", "$158.50", "←", "USD $1,000", "NT$31,520"],
        ["NT$10,000", "$317.00", "←", "USD $10,000", "NT$315,200"],
    ], [16, 14, 6, 14, 16])
    
    # ===== Sheet 3: 台幣 <-> 日圓 =====
    ws3 = wb.create_sheet("台幣 <-> 日圓")
    ws3.merge_cells('A1:E1')
    ws3['A1'] = '台幣 (TWD) 兌日圓 (JPY) 換算對照表'
    ws3['A1'].font = title_font
    
    write_header(ws3, 3, ["台幣金額", "日圓金額", "", "日圓金額", "台幣金額"])
    write_data(ws3, 4, [
        ["NT$100", "¥508", "←", "JPY ¥1,000", "NT$197"],
        ["NT$500", "¥2,540", "←", "JPY ¥5,000", "NT$984"],
        ["NT$1,000", "¥5,080", "←", "JPY ¥10,000", "NT$1,969"],
        ["NT$5,000", "¥25,400", "←", "JPY ¥50,000", "NT$9,843"],
        ["NT$10,000", "¥50,800", "←", "JPY ¥100,000", "NT$19,685"],
    ], [16, 14, 6, 14, 16])
    
    # ===== Sheet 4: 市場概況 =====
    ws4 = wb.create_sheet("市場概況")
    ws4.merge_cells('A1:B1')
    ws4['A1'] = '近期市場狀況與注意事項'
    ws4['A1'].font = title_font
    
    write_header(ws4, 3, ["指標", "數據"])
    write_data(ws4, 4, [
        ["📅 數據日期", "2026 年 6 月 5 日"],
        ["📈 USD/TWD 走勢", "區間 31.47～31.52，波動平穩"],
        ["📈 TWD/JPY 走勢", "區間 ¥5.07～¥5.09，小幅波動"],
        ["📊 台幣 12 個月趨勢", "相對美元貶值 5.31%"],
        ["📊 日圓 12 個月趨勢", "相對台幣貶值 6.12%"],
        ["⚠️ 報價說明", "以上為國際中間價 (mid-market rate)"],
        ["🏦 銀行價差", "實際換匯價差通常約 1～3%"],
        ["🌙 交易時間", "週一至週五 8:00～17:00"],
        ["📲 查詢建議", "出發前請至銀行官網查最新報價"],
    ], [22, 42])
    
    # 儲存檔案
    output_path = 'TWD_Exchange_Rates_20260605.xlsx'
    wb.save(output_path)
    
    # 驗證與輸出
    print("=" * 50)
    print("✅ Excel 檔案產生成功！")
    print("=" * 50)
    print(f"📁 路徑        : {output_path}")
    print(f"📊 檔案大小    : {os.path.getsize(output_path):,} bytes "
          f"({os.path.getsize(output_path) / 1024:.1f} KB)")
    print(f"✅ 檔案存在    : {os.path.exists(output_path)}")
    print(f"📋 工作表數量 : {len(wb.sheetnames)}")
    print(f"📋 工作表列表 :")
    for s in wb.sheetnames:
        print(f"   - {s}")
    print("=" * 50)
    
    return output_path


if __name__ == "__main__":
    result = create_rates_excel()
    print(f"\nFinal result: {result}")