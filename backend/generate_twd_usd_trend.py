import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
import os

print("=" * 60)
print("正在生成台幣兌美元匯率趨勢 Excel 檔案...")
print("=" * 60)

# 讀取數據
df = pd.read_csv('twd_usd_data.csv')
print(f"\n📊 數據筆數：{len(df)}")
print(f"\n📋 數據內容：")
print(df.to_string(index=False))

# 建立工作簿
wb = openpyxl.Workbook()

# 定義樣式
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(name='Microsoft JhengHei', bold=True, color='FFFFFF', size=11)
title_font = Font(name='Microsoft JhengHei', bold=True, size=14, color='1F4E79')
subtitle_font = Font(name='Microsoft JhengHei', bold=True, size=12, color='1F4E79')
data_font = Font(name='Microsoft JhengHei', size=10)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 日期格式化函數
def format_date(date_str):
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return dt.strftime('%Y/%m/%d')
    except:
        return date_str

# 匯率格式化函數
def format_rate(rate):
    try:
        return round(float(rate), 6)
    except:
        return rate

# ==========================================
# Sheet 1: 匯率趨勢圖表
# ==========================================
ws1 = wb.active
ws1.title = "匯率趨勢圖"
ws1.sheet_properties.tabColor = "1F4E79"

# 標題
ws1.merge_cells('A1:F1')
ws1['A1'] = '台幣 (TWD) 兌美元 (USD) 近一周匯率趨勢圖'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[1].height = 35

# 副標題
ws1.merge_cells('A2:F2')
ws1['A2'] = f'數據期間：2026/05/28 ~ 2026/06/05 | 數據來源：Yahoo Finance'
ws1['A2'].font = Font(name='Microsoft JhengHei', size=10, color='666666')
ws1['A2'].alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[2].height = 25

# 寫入數據標題列
row_num = 4
headers = ['日期', '開盤', '最高', '最低', '收盤', '日變化']
for col, val in enumerate(headers, 1):
    cell = ws1.cell(row=row_num, column=col, value=val)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style

# 寫入數據行
current_close = None
for idx, (_, row) in enumerate(df.iterrows(), start=1):
    date_str = row['日期']
    open_rate = row['開盤']
    high_rate = row['最高']
    low_rate = row['最低']
    close_rate = row['收盤']
    
    # 計算日變化
    if current_close is not None and close_rate is not None and current_close is not None:
        change = close_rate - current_close
        change_pct = (change / current_close) * 100
    else:
        change = 0
        change_pct = 0
    
    current_close = close_rate
    
    # 寫入數據
    ws1.cell(row=row_num + idx, column=1, value=format_date(date_str)).font = data_font
    ws1.cell(row=row_num + idx, column=1).alignment = data_align
    ws1.cell(row=row_num + idx, column=1).border = border_style
    
    ws1.cell(row=row_num + idx, column=2, value=format_rate(open_rate)).font = data_font
    ws1.cell(row=row_num + idx, column=2).alignment = data_align
    ws1.cell(row=row_num + idx, column=2).border = border_style
    
    ws1.cell(row=row_num + idx, column=3, value=format_rate(high_rate)).font = data_font
    ws1.cell(row=row_num + idx, column=3).alignment = data_align
    ws1.cell(row=row_num + idx, column=3).border = border_style
    
    ws1.cell(row=row_num + idx, column=4, value=format_rate(low_rate)).font = data_font
    ws1.cell(row=row_num + idx, column=4).alignment = data_align
    ws1.cell(row=row_num + idx, column=4).border = border_style
    
    ws1.cell(row=row_num + idx, column=5, value=format_rate(close_rate)).font = data_font
    ws1.cell(row=row_num + idx, column=5).alignment = data_align
    ws1.cell(row=row_num + idx, column=5).border = border_style
    
    # 變化率格式化
    change_str = f"{change:+.6f}"
    change_pct_str = f"{change_pct:+.2f}%"
    ws1.cell(row=row_num + idx, column=6, value=change_pct_str).font = data_font
    ws1.cell(row=row_num + idx, column=6).alignment = data_align
    ws1.cell(row=row_num + idx, column=6).border = border_style
    
    # 顏色標記
    if change_pct > 0:
        ws1.cell(row=row_num + idx, column=6).font = Font(name='Microsoft JhengHei', size=10, color='00B050', bold=True)
    elif change_pct < 0:
        ws1.cell(row=row_num + idx, column=6).font = Font(name='Microsoft JhengHei', size=10, color='FF0000', bold=True)

# 設置列寬
col_widths = [14, 14, 14, 14, 14, 14]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 設置行高
for i in range(row_num, row_num + len(df) + 2):
    ws1.row_dimensions[i].height = 25

# 建立折線圖
chart = LineChart()
chart.type = "line"
chart.title = "TWD/USD 匯率趨勢 (近一周)"
chart.y_axis.title = "匯率 (TWD/USD)"
chart.x_axis.title = "日期"
chart.style = 10
chart.width = 28
chart.height = 16

# 數據範圍
dates = Reference(ws1, min_col=1, min_row=row_num + 1, max_row=row_num + len(df))
open_data = Reference(ws1, min_col=2, min_row=row_num, max_row=row_num + len(df))
high_data = Reference(ws1, min_col=3, min_row=row_num, max_row=row_num + len(df))
low_data = Reference(ws1, min_col=4, min_row=row_num, max_row=row_num + len(df))
close_data = Reference(ws1, min_col=5, min_row=row_num, max_row=row_num + len(df))

chart.add_data(open_data, titles_from_data=True)
chart.add_data(high_data, titles_from_data=True)
chart.add_data(low_data, titles_from_data=True)
chart.add_data(close_data, titles_from_data=True)

chart.set_categories(dates)

# 設置圖例
chart.legend.position = 'bottom'
chart.legend.include_legend_in_print = True

# 設置樣式
chart.series[0].name = '開盤'
chart.series[1].name = '最高'
chart.series[2].name = '最低'
chart.series[3].name = '收盤'

chart.series[0].graphicalProperties.line.color.rgb = '1F4E79'
chart.series[1].graphicalProperties.line.color.rgb = '00B050'
chart.series[2].graphicalProperties.line.color.rgb = 'FF0000'
chart.series[3].graphicalProperties.line.color.rgb = 'FFC000'

# 添加圖表到工作表
ws1.add_chart(chart, "A15")

# ==========================================
# Sheet 2: 數據總覽
# ==========================================
ws2 = wb.create_sheet("數據總覽")
ws2.sheet_properties.tabColor = "00B050"

# 標題
ws2.merge_cells('A1:F1')
ws2['A1'] = '台幣 (TWD) 兌美元 (USD) 近一周匯率數據總覽'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws2.row_dimensions[1].height = 35

# 統計信息
stats_row = 3
ws2.cell(row=stats_row, column=1, value="統計期間").font = subtitle_font
ws2.cell(row=stats_row, column=2, value="2026/05/28 ~ 2026/06/05").font = data_font
ws2.cell(row=stats_row, column=1).border = border_style
ws2.cell(row=stats_row, column=2).border = border_style

ws2.cell(row=stats_row + 1, column=1, value="最高收盤價").font = subtitle_font
ws2.cell(row=stats_row + 1, column=2, value=max(df['收盤'])).font = data_font
ws2.cell(row=stats_row + 1, column=1).border = border_style
ws2.cell(row=stats_row + 1, column=2).border = border_style

ws2.cell(row=stats_row + 2, column=1, value="最低收盤價").font = subtitle_font
ws2.cell(row=stats_row + 2, column=2, value=min(df['收盤'])).font = data_font
ws2.cell(row=stats_row + 2, column=1).border = border_style
ws2.cell(row=stats_row + 2, column=2).border = border_style

ws2.cell(row=stats_row + 3, column=1, value="平均收盤價").font = subtitle_font
ws2.cell(row=stats_row + 3, column=2, value=df['收盤'].mean()).font = data_font
ws2.cell(row=stats_row + 3, column=1).border = border_style
ws2.cell(row=stats_row + 3, column=2).border = border_style

ws2.cell(row=stats_row + 4, column=1, value="數據來源").font = subtitle_font
ws2.cell(row=stats_row + 4, column=2, value="Yahoo Finance").font = data_font
ws2.cell(row=stats_row + 4, column=1).border = border_style
ws2.cell(row=stats_row + 4, column=2).border = border_style

# 詳細數據表
ws2.merge_cells('A6:F6')
ws2['A6'] = '詳細數據'
ws2['A6'].font = subtitle_font

headers2 = ['日期', '開盤', '最高', '最低', '收盤']
for col, val in enumerate(headers2, 1):
    cell = ws2.cell(row=7, column=col, value=val)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style

for idx, (_, row) in enumerate(df.iterrows(), start=8):
    ws2.cell(row=idx, column=1, value=format_date(row['日期'])).font = data_font
    ws2.cell(row=idx, column=1).alignment = data_align
    ws2.cell(row=idx, column=1).border = border_style
    
    for col in range(2, 6):
        ws2.cell(row=idx, column=col, value=format_rate(row[headers2[col-1]])).font = data_font
        ws2.cell(row=idx, column=col).alignment = data_align
        ws2.cell(row=idx, column=col).border = border_style

col_widths2 = [14, 14, 14, 14, 14]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ==========================================
# Sheet 3: 趨勢分析
# ==========================================
ws3 = wb.create_sheet("趨勢分析")
ws3.sheet_properties.tabColor = "FFC000"

# 標題
ws3.merge_cells('A1:C1')
ws3['A1'] = '台幣 (TWD) 兌美元 (USD) 趨勢分析'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws3.row_dimensions[1].height = 35

# 趨勢分析數據
analysis_headers = ['指標', '數值', '說明']
for col, val in enumerate(analysis_headers, 1):
    cell = ws3.cell(row=3, column=col, value=val)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style

analysis_data = [
    ["最高收盤價", f"{max(df['收盤']):.6f}", "一周內最高收盤價格"],
    ["最低收盤價", f"{min(df['收盤']):.6f}", "一周內最低收盤價格"],
    ["平均收盤價", f"{df['收盤'].mean():.6f}", "一周平均收盤價格"],
    ["收盤價標準差", f"{df['收盤'].std():.6f}", "價格波動程度"],
    ["漲跌天數", f"{(df['收盤'] > df['收盤'].shift(1)).sum()}", "相比前一日上漲的天數"],
    ["跌跌天數", f"{(df['收盤'] < df['收盤'].shift(1)).sum()}", "相比前一日下跌的天數"],
    ["最大漲幅", f"{(df['收盤'].pct_change().max()*100):.2f}%", "單日最大漲幅"],
    ["最大跌幅", f"{(df['收盤'].pct_change().min()*100):.2f}%", "單日最大跌幅"],
]

for idx, (indicator, value, desc) in enumerate(analysis_data, start=4):
    ws3.cell(row=idx, column=1, value=indicator).font = data_font
    ws3.cell(row=idx, column=1).alignment = left_align
    ws3.cell(row=idx, column=1).border = border_style
    
    ws3.cell(row=idx, column=2, value=value).font = data_font
    ws3.cell(row=idx, column=2).alignment = data_align
    ws3.cell(row=idx, column=2).border = border_style
    
    ws3.cell(row=idx, column=3, value=desc).font = data_font
    ws3.cell(row=idx, column=3).alignment = left_align
    ws3.cell(row=idx, column=3).border = border_style

col_widths3 = [16, 16, 30]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# 儲存檔案
output_path = 'TWD_USD_Trend_Chart_20260605.xlsx'
wb.save(output_path)

# 驗證
print("\n" + "=" * 60)
print("✅ Excel 檔案產生成功！")
print("=" * 60)
print(f"📁 路徑        : {output_path}")
print(f"📊 檔案大小    : {os.path.getsize(output_path):,} bytes "
      f"({os.path.getsize(output_path) / 1024:.1f} KB)")
print(f"✅ 檔案存在    : {os.path.exists(output_path)}")
print(f"📋 工作表數量 : {len(wb.sheetnames)}")
print(f"📋 工作表列表 :")
for s in wb.sheetnames:
    print(f"   - {s}")
print("=" * 60)

print("\n✅ 完成！")